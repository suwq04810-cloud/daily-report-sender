# -*- coding: utf-8 -*-
import base64
from datetime import datetime
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
import os
import smtplib
import time
import openpyxl


# =========================================================================
# 0. 提取 Logo 图片转为 Base64 字符串
# =========================================================================
def get_logo_base64_html():
  logo_file = "logo.png"
  if os.path.exists(logo_file):
    try:
      with open(logo_file, "rb") as f:
        b64_data = base64.b64encode(f.read()).decode("utf-8")
        return f'<div style="margin-top: 10px;"><img src="data:image/png;base64,{b64_data}" style="max-width: 320px; height: auto; display: block;" alt="iSoftStone Logo" /></div>'
    except Exception as e:
      print(f"读取 logo 图片失败: {e}")
  return ""


# =========================================================================
# 通用签名模块（完美还原 Foxmail 签名样式）
# =========================================================================
def get_signature_html():
  logo_html = get_logo_base64_html()
  return f"""
    <div style="margin-top: 25px; font-family: 'Microsoft YaHei UI', 'Microsoft YaHei', 'Verdana', sans-serif; font-size: 10pt; color: #333333; line-height: 1.6; border-top: 1px solid #CCCCCC; padding-top: 12px; width: 360px;">
        <div style="font-weight: bold; font-size: 11pt; color: #000000; margin-bottom: 2px;">苏文强</div>
        <div>呼市HW中国区政企平软上云实施部</div>
        <div>软通动力技术服务有限公司</div>
        <div><span style="color: #FF0000; font-weight: bold;">手机：</span>+86 16611601206</div>
        <div><span style="color: #FF0000; font-weight: bold;">邮箱：</span><a href="mailto:wqsud@isoftstone.com" style="color: #000000; text-decoration: underline;">wqsud@isoftstone.com</a></div>
        <div><a href="http://www.isoftstone.com" style="color: #0066CC; text-decoration: underline;">www.isoftstone.com</a></div>
        {logo_html}
    </div>
    """


# =========================================================================
# 1. 日报 HTML 生成（提取 '1.日工作简报'）
# =========================================================================
def generate_daily_html(wb):
  sheet = wb["1.日工作简报"] if "1.日工作简报" in wb.sheetnames else wb.active

  title = (
      str(sheet["A1"].value).strip()
      if sheet["A1"].value
      else f'工作&学习日报-{datetime.now().strftime("%m月%d日")}'
  )
  today_work = str(sheet["A3"].value or "").strip().replace("\n", "<br>")
  tomorrow_plan = str(sheet["A5"].value or "").strip().replace("\n", "<br>")
  sig_html = get_signature_html()

  html_content = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
</head>
<body style="margin: 0; padding: 10px; background-color: #FFFFFF;">
    <table style="border-collapse: collapse; width: 100%; max-width: 800px; border: 1px solid #000000; table-layout: fixed; box-sizing: border-box; font-family: SimSun, '宋体', 'Microsoft YaHei', '微软雅黑', sans-serif;">
        <!-- 1. 标题行 -->
        <tr>
            <td style="border: 1px solid #000000; background-color: #8DB4E2; font-family: 'SimSun', '宋体', serif; font-size: 18pt; font-weight: bold; color: #000000; text-align: center; vertical-align: middle; height: 46px; padding: 6px 10px;">
                {title}
            </td>
        </tr>

        <!-- 2. 今日工作概述 栏目头 -->
        <tr>
            <td style="border: 1px solid #000000; background-color: #DBE5F1; text-align: left; vertical-align: middle; height: 30px; padding: 4px 10px;">
                <span style="font-family: 'SimSun', '宋体', serif; font-size: 12pt; font-weight: bold; color: #000000;">今日工作概述</span>
                <span style="font-family: 'SimSun', '宋体', serif; font-size: 10pt; font-weight: bold; color: #FF0000;">（重点工作、重大风险说明(加粗标红）、周边求助(加粗标红），小于3条）</span>
            </td>
        </tr>

        <!-- 3. 今日工作内容 正文 -->
        <tr>
            <td style="border: 1px solid #000000; background-color: #FFFFFF; font-family: 'STXihei', '华文细黑', 'Microsoft YaHei', '微软雅黑', sans-serif; font-size: 10pt; font-weight: bold; color: #000000; text-align: left; vertical-align: top; padding: 12px 10px; white-space: pre-wrap; line-height: 1.6; min-height: 120px; height: 120px;">
                {today_work}
            </td>
        </tr>

        <!-- 4. 明日工作计划 栏目头 -->
        <tr>
            <td style="border: 1px solid #000000; background-color: #DBE5F1; text-align: left; vertical-align: middle; height: 30px; padding: 4px 10px;">
                <span style="font-family: 'SimSun', '宋体', serif; font-size: 12pt; font-weight: bold; color: #000000;">明日工作计划</span>
                <span style="font-family: 'SimSun', '宋体', serif; font-size: 10pt; font-weight: bold; color: #FF0000;">（重点工作、重大风险说明(加粗标红）、周边求助(加粗标红），小于3条）</span>
            </td>
        </tr>

        <!-- 5. 明日工作计划 正文 -->
        <tr>
            <td style="border: 1px solid #000000; background-color: #FFFFFF; font-family: 'STXihei', '华文细黑', 'Microsoft YaHei', '微软雅黑', sans-serif; font-size: 10pt; font-weight: bold; color: #000000; text-align: left; vertical-align: top; padding: 12px 10px; white-space: pre-wrap; line-height: 1.6; min-height: 90px; height: 90px;">
                {tomorrow_plan}
            </td>
        </tr>

        <!-- 6. 底部装饰条 -->
        <tr>
            <td style="border: 1px solid #000000; background-color: #DBE5F1; height: 18px; padding: 0;"></td>
        </tr>
    </table>
    
    <!-- 底部专属签名 -->
    {sig_html}
</body>
</html>"""
  return title, html_content


# =========================================================================
# 2. 周报 HTML 生成
# =========================================================================
def generate_weekly_html(wb):
  sheet = wb["周工作简报"] if "周工作简报" in wb.sheetnames else wb.active

  r1_title = str(sheet["A1"].value or "能力提升-学习周报").strip()
  curr_prod = str(sheet["C2"].value or "云计算&中级").strip()
  target_prod = str(sheet["G2"].value or "云计算&高级").strip()
  report_period = (
      str(sheet["J2"].value or "2026年8月24日-2026年8月28日").strip()
  )
  plan_period = str(sheet["C3"].value or "20周").strip()
  user_name = str(sheet["F4"].value or "苏文强").strip()
  user_phone = str(sheet["H4"].value or "16611601206").strip()

  this_week_work = str(sheet["A6"].value or "").strip().replace("\n", "<br>")
  next_week_plan = (
      str(sheet["A8"].value or "学习华为私有云部署").strip().replace("\n", "<br>")
  )

  # 标题带上动态日期，彻底防止邮件折叠
  subject = f"苏文强学习周报-{datetime.now().strftime('%m月%d日')}"

  red_slash = '<span style="color: #FF0000; font-weight: bold;">/</span>'

  style_main_title = "border: 1px solid #000000; background-color: #CCCCFF; color: #0000FF; font-family: 'Microsoft YaHei', '微软雅黑', 'SimHei', '黑体', sans-serif; font-size: 20pt; font-weight: 900; text-align: center; vertical-align: middle; height: 48px; padding: 6px 10px; letter-spacing: 2px;"
  style_section_bar = "border: 1px solid #000000; background-color: #CCCCFF; color: #0000FF; font-family: 'Microsoft YaHei', '微软雅黑', 'SimHei', '黑体', sans-serif; font-size: 14pt; font-weight: 900; text-align: center; vertical-align: middle; height: 36px; padding: 4px 8px; letter-spacing: 1px;"
  style_info_bold = "border: 1px solid #000000; background-color: #FFFFFF; font-family: 'Microsoft YaHei', '微软雅黑', sans-serif; font-size: 10pt; font-weight: bold; color: #000000; text-align: center; vertical-align: middle; padding: 4px 6px;"
  style_info_val = "border: 1px solid #000000; background-color: #FFFFFF; font-family: 'Microsoft YaHei', '微软雅黑', sans-serif; font-size: 10pt; color: #000000; text-align: center; vertical-align: middle; padding: 4px 6px;"
  style_content_yellow = "border: 1px solid #000000; background-color: #FFFF99; color: #000000; font-family: 'Microsoft YaHei', '微软雅黑', 'SimSun', '宋体', sans-serif; font-size: 10.5pt; line-height: 1.6; text-align: left !important; vertical-align: top; padding: 12px 14px; white-space: pre-wrap;"
  style_header_yellow = "border: 1px solid #000000; background-color: #FFFF99; color: #000000; font-family: 'Microsoft YaHei', '微软雅黑', sans-serif; font-size: 10pt; font-weight: bold; text-align: center; vertical-align: middle; padding: 4px 6px;"
  style_val_yellow = "border: 1px solid #000000; background-color: #FFFF99; color: #000000; font-family: 'Microsoft YaHei', '微软雅黑', sans-serif; font-size: 10pt; text-align: center; vertical-align: middle; padding: 4px 6px;"

  sig_html = get_signature_html()

  html_content = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
</head>
<body style="margin: 0; padding: 10px; background-color: #FFFFFF;">
    <table style="border-collapse: collapse; width: 100%; max-width: 900px; border: 1px solid #000000; table-layout: fixed; box-sizing: border-box; font-family: 'Microsoft YaHei', '微软雅黑', 'SimSun', '宋体', sans-serif;">
        <tr>
            <td colspan="8" style="{style_main_title}">{r1_title}</td>
        </tr>
        <tr>
            <td style="{style_info_bold}">当前产品&级别</td>
            <td style="{style_info_val}">{curr_prod}</td>
            <td style="{style_info_bold}">目标产品&级别</td>
            <td style="{style_info_val}">{target_prod}</td>
            <td style="{style_info_bold}">报告周期</td>
            <td colspan="3" style="{style_info_val} font-size: 9pt;">{report_period}</td>
        </tr>
        <tr>
            <td style="{style_info_bold}">计划周期</td>
            <td colspan="3" style="{style_info_val}">{plan_period}</td>
            <td style="{style_info_bold}">姓名</td>
            <td style="{style_info_val}">{user_name}</td>
            <td style="{style_info_bold}">电话</td>
            <td style="{style_info_val}">{user_phone}</td>
        </tr>
        <tr>
            <td colspan="4" style="{style_info_bold}">交付工程师信息</td>
            <td colspan="2" style="{style_info_val}">{user_name}</td>
            <td colspan="2" style="{style_info_val}">{user_phone}</td>
        </tr>
        <tr>
            <td colspan="8" style="{style_section_bar}">本周工作内容</td>
        </tr>
        <tr>
            <td colspan="8" style="{style_content_yellow}">{this_week_work}</td>
        </tr>
        <tr>
            <td colspan="8" style="{style_section_bar}">下周工作计划</td>
        </tr>
        <tr>
            <td colspan="8" style="{style_content_yellow} font-weight: bold; min-height: 50px;">{next_week_plan}</td>
        </tr>
        <tr>
            <td colspan="8" style="{style_section_bar}">项目风险问题</td>
        </tr>
        <tr>
            <td style="{style_header_yellow} width: 7%;">序号</td>
            <td style="{style_header_yellow} width: 12%; color: #FF0000;">产生日期</td>
            <td style="{style_header_yellow} width: 25%;">问题描述&影响</td>
            <td style="{style_header_yellow} width: 10%;">紧急程度</td>
            <td style="{style_header_yellow} width: 20%;">规避措施、解决进展</td>
            <td style="{style_header_yellow} width: 8%;">责任人</td>
            <td style="{style_header_yellow} width: 8%;">状态</td>
            <td style="{style_header_yellow} width: 10%;">实际关闭日期</td>
        </tr>
        <tr>
            <td style="{style_val_yellow}">1</td>
            <td style="{style_val_yellow}">{red_slash}</td>
            <td style="{style_val_yellow}">{red_slash}</td>
            <td style="{style_val_yellow}">{red_slash}</td>
            <td style="{style_val_yellow}">{red_slash}</td>
            <td style="{style_val_yellow}">{red_slash}</td>
            <td style="{style_val_yellow}">{red_slash}</td>
            <td style="{style_val_yellow}">{red_slash}</td>
        </tr>
        <tr>
            <td colspan="8" style="{style_section_bar}">求助</td>
        </tr>
        <tr>
            <td style="{style_header_yellow} width: 7%;">序号</td>
            <td colspan="3" style="{style_header_yellow}">求助事宜</td>
            <td style="{style_header_yellow} width: 12%;">求助人</td>
            <td style="{style_header_yellow} width: 16%;">要求解决时间</td>
            <td colspan="2" style="{style_header_yellow}">备注</td>
        </tr>
        <tr>
            <td style="{style_val_yellow}">1</td>
            <td colspan="3" style="{style_val_yellow}">{red_slash}</td>
            <td style="{style_val_yellow}">{red_slash}</td>
            <td style="{style_val_yellow}">{red_slash}</td>
            <td colspan="2" style="{style_val_yellow}">{red_slash}</td>
        </tr>
    </table>

    <!-- 底部专属签名 -->
    {sig_html}
</body>
</html>"""
  return subject, html_content


# =========================================================================
# 3. 发送邮件底层函数
# =========================================================================
def send_email_message(subject, html_body, to_list, cc_list):
  mail_user = os.environ.get("MAIL_USER")
  mail_pass = os.environ.get("MAIL_PASS")
  mail_host = os.environ.get("MAIL_HOST", "smtp.isoftstone.com")

  if not mail_user or not mail_pass:
    raise ValueError(
        "❌ 错误: 缺少 MAIL_USER 或 MAIL_PASS 环境变量，请检查 GitHub Secrets"
        " 配置！"
    )

  msg = MIMEMultipart()
  msg["From"] = f"wqsud <{mail_user}>"
  msg["To"] = ",".join(to_list)
  if cc_list:
    msg["Cc"] = ",".join(cc_list)
  msg["Subject"] = Header(subject, "utf-8")
  msg["Date"] = formatdate(localtime=True)  # 标准时间戳
  msg.attach(MIMEText(html_body, "html", "utf-8"))

  server = smtplib.SMTP_SSL(mail_host, 465, timeout=30)
  server.login(mail_user, mail_pass)
  server.sendmail(mail_user, to_list + cc_list, msg.as_string())
  server.quit()
  print(f"✅ 邮件投递成功 -> 主题:「{subject}」 | 收件人: {to_list}")


# =========================================================================
# 4. 主流程调度
# =========================================================================
def main():
  excel_file = "daily_report.xlsx"
  if not os.path.exists(excel_file):
    print(f"❌ 未找到 {excel_file} 文件，跳过本次发送。")
    return

  wb = openpyxl.load_workbook(excel_file, data_only=True)

  # =========================================================================
  # 【收发模式配置】
  # =========================================================================

  # 👉 模式一：【测试模式】（测试时所有邮件只发给自己）
  # daily_to = ["wqsud@isoftstone.com"]
  # daily_cc = []
  # weekly_to = ["wqsud@isoftstone.com"]
  # weekly_cc = []

  # 👉 模式二：【正式发送模式】（正式上线时启用）
  daily_to = ["hdliuf@isoftstone.com"]
  daily_cc = ["weiliuay@isoftstone.com", "zycaoc@isoftstone.com", "nawangam@isoftstone.com"]
  weekly_to = ["zycaoc@isoftstone.com"]
  weekly_cc = []

  # =========================================================================

  # 1. 发送【日报】
  print("\n==============================================")
  print(" [1/2] 正在投递：今日工作日报 ...")
  daily_subj, daily_html = generate_daily_html(wb)
  send_email_message(daily_subj, daily_html, daily_to, daily_cc)
  print("==============================================")

  # 2. 判断周五并发送【周报】
  # 提示：周五时 weekday() 为 4
  is_friday = datetime.now().weekday() == 4

  if is_friday:
    print("\n⏳ 等待 4 秒，确保邮件服务器安全接收下一封...")
    time.sleep(4)

    print("==============================================")
    print(" [2/2] 正在投递：学习周报 ...")
    weekly_subj, weekly_html = generate_weekly_html(wb)
    send_email_message(weekly_subj, weekly_html, weekly_to, weekly_cc)
    print("==============================================")
  else:
    print("今日非周五，跳过周报。")


if __name__ == "__main__":
  main()
